from fastapi import FastAPI
from fastapi.responses import FileResponse, HTMLResponse
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import os
from datetime import datetime

app = FastAPI(title="Axiora Portfolio - Reportes Financieros")

# Configuración de BD

DATA_DIR = 'data'
DB_PATH = os.path.join(DATA_DIR, 'financiera.db')
engine = create_engine(f'sqlite:///{DB_PATH}')

@app.get("/")
async def root():
    return HTMLResponse("""
    <h1>Portafolio Python - Luis Melendez</h1>
    <p><a href='/docs'>Ver Documentación Interactiva (Swagger)</a></p>
    <p><a href='/reporte'>Generar y Descargar Reporte Excel</a></p>
    """)

@app.get("/reporte")
async def generar_reporte():
    
    # Extraer y transformar datos

    df = pd.read_sql("SELECT * FROM operaciones", engine)
    df['fecha'] = pd.to_datetime(df['fecha'])
    df['monto'] = pd.to_numeric(df['monto'])
    df['mes'] = df['fecha'].dt.strftime('%Y-%m')
    
    resumen = df.groupby('producto').agg({
        'monto': ['sum', 'mean', 'count'],
        'estado': lambda x: (x == 'Aprobado').sum()
    }).round(2)
    resumen.columns = ['Monto_Total', 'Monto_Promedio', 'Cantidad', 'Aprobados']

    # Generar Excel

    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
    filename = f'Reporte_Financiero_{fecha_hoy}.xlsx'
    filepath = os.path.join('output', filename)
    os.makedirs('output', exist_ok=True)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Datos', index=False)
        resumen.to_excel(writer, sheet_name='Resumen')

    wb = load_workbook(filepath)

    # Hoja Datos

    ws = wb['Datos']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if isinstance(cell.value, datetime):
                        length = 12
                    if length > max_length:
                        max_length = length
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 5, 25)
    
    # Hoja Resumen

    ws_resumen = wb['Resumen']
    for column in ws_resumen.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_resumen.column_dimensions[column_letter].width = min(max_length + 5, 20)
    
    # Gráfico

    chart = BarChart()
    chart.type = 'col'
    chart.style = 10
    chart.title = 'Monto por Producto'
    chart.y_axis.title = 'Monto (S/)'
    chart.x_axis.title = 'Producto'
    
    data = Reference(ws_resumen, min_col=2, min_row=1, max_row=5, max_col=2)
    cats = Reference(ws_resumen, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_resumen.add_chart(chart, 'G2')
    
    wb.save(filepath)
    
    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
