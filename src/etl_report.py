import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from sqlalchemy import create_engine, text
from datetime import datetime
import os

# ====================== CONFIGURACIÓN ======================

OUTPUT_DIR = 'output'
DATA_DIR = 'data'
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, 'financiera.db')
engine = create_engine(f'sqlite:///{DB_PATH}')

# ====================== CREAR TABLA Y DATOS ======================

def inicializar_base_datos():
    with engine.connect() as conn:
        # Crear tabla
        
        conn.execute(text('''
            CREATE TABLE IF NOT EXISTS operaciones (
                id INTEGER PRIMARY KEY,
                fecha TEXT,
                cliente TEXT,
                monto REAL,
                producto TEXT,
                estado TEXT
            )
        '''))
        
        # Verificar si ya hay datos

        result = conn.execute(text('SELECT COUNT(*) FROM operaciones')).scalar()
        if result == 0:

            # Insertar datos usando lista de diccionarios

            datos = [
                {'fecha': '2026-06-01', 'cliente': 'Cliente A', 'monto': 1250.50, 'producto': 'Préstamo Personal', 'estado': 'Aprobado'},
                {'fecha': '2026-06-02', 'cliente': 'Cliente B', 'monto': 980.75, 'producto': 'Tarjeta Crédito', 'estado': 'Pendiente'},
                {'fecha': '2026-06-03', 'cliente': 'Cliente A', 'monto': 2340.00, 'producto': 'Hipoteca', 'estado': 'Aprobado'},
                {'fecha': '2026-06-04', 'cliente': 'Cliente C', 'monto': 675.25, 'producto': 'Préstamo Personal', 'estado': 'Rechazado'},
                {'fecha': '2026-06-05', 'cliente': 'Cliente B', 'monto': 1540.80, 'producto': 'Tarjeta Crédito', 'estado': 'Aprobado'},
                {'fecha': '2026-06-06', 'cliente': 'Cliente D', 'monto': 3200.00, 'producto': 'Hipoteca', 'estado': 'Aprobado'}
            ]
            
            conn.execute(
                text('INSERT INTO operaciones (fecha, cliente, monto, producto, estado) VALUES (:fecha, :cliente, :monto, :producto, :estado)'),
                datos
            )
            conn.commit()
            print('Datos insertados correctamente')
        else:
            print('La base de datos ya tiene datos')

# ====================== EXTRACT ======================

def extraer_datos():
    query = "SELECT * FROM operaciones"
    df = pd.read_sql(query, engine)
    print(f'Datos extraídos: {len(df)} registros')
    return df

# ====================== TRANSFORM ======================

def transformar_datos(df):
    df['fecha'] = pd.to_datetime(df['fecha'])
    df['monto'] = pd.to_numeric(df['monto'])
    df['mes'] = df['fecha'].dt.strftime('%Y-%m')
    
    resumen = df.groupby('producto').agg({
        'monto': ['sum', 'mean', 'count'],
        'estado': lambda x: (x == 'Aprobado').sum()
    }).round(2)
    resumen.columns = ['Monto_Total', 'Monto_Promedio', 'Cantidad', 'Aprobados']
    
    print('Transformación completada')
    return df, resumen

# ====================== GENERAR REPORTE ======================

def generar_reporte_excel(df, resumen):
    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
    filename = f'Reporte_Financiero_{fecha_hoy}.xlsx'
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Datos', index=False)
        resumen.to_excel(writer, sheet_name='Resumen')
    
    wb = load_workbook(filepath)
    ws = wb['Datos']
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 25)
    
    # Gráfico

    ws_resumen = wb['Resumen']
    chart = BarChart()
    chart.type = 'col'
    chart.style = 10
    chart.title = 'Monto por Producto'
    chart.y_axis.title = 'Monto (S/)'
    chart.x_axis.title = 'Producto'
    
    data = Reference(ws_resumen, min_col=2, min_row=1, max_row=5, max_col=2)
    cats = Reference(ws_resumen, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_resumen.add_chart(chart, 'G2')
    
    wb.save(filepath)
    print(f'Reporte generado: {filepath}')
    return filepath

# ====================== EJECUCIÓN ======================

if __name__ == '__main__':
    inicializar_base_datos()
    df = extraer_datos()
    df, resumen = transformar_datos(df)
    generar_reporte_excel(df, resumen)
    print('\n¡Proceso ETL + Reporte completado exitosamente!')
